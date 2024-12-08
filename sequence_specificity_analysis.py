import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

"""define amino acid groups"""
AMINO_ACID_GROUPS = {
    "nonpolar": {'A', 'V', 'L', 'C', 'I', 'M', 'F', 'P', 'W', 'G'},
    "polar uncharged": {'S', 'T', 'N', 'Q', 'Y'},
    "positive": {'K', 'R', 'H'},
    "negative": {'D', 'E'},
}


"""determine the group of an amino acid"""
def get_amino_acid_group(aa):
    for group, amino_acids in AMINO_ACID_GROUPS.items():
        if aa in amino_acids:
            return group
    return None


"""analyze mutations in compare_positions_df, determine for every position if it's a radical or neutral change"""
def analyze_mutations(df, comparison_type, taxonomy, num_seq):
    mutation_results = {}

    for position in df.columns:
        residues = df[position]  # Get all residues for the current position
        groups = {get_amino_acid_group(aa) for aa in residues if pd.notna(aa)}  # Unique groups

        # Determine mutation type
        if len(groups) > 1:
            mutation_results[position] = str(groups)
        else:
            mutation_results[position] = None

    # Add mutation type as a new row
    mutation_row = pd.Series(mutation_results, name="Mutation_Type")
    compare_positions_df = pd.concat([df, pd.DataFrame([mutation_row])])
    result_df = find_mutant_segments(compare_positions_df)
    last_row = len(result_df)

    file_name = f"{num_seq}seq_{taxonomy}_{comparison_type}_compare_and_analyze.xlsx"
    result_df.to_excel(file_name)
    wb = load_workbook(file_name)
    ws = wb.active

    fill1 = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for cell in ws[last_row]:
        if cell.value is not None:
            cell.fill = fill1

    fill2 = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
    for cell in ws[last_row+1]:
        if cell.value == "*":
            cell.fill = fill2

    wb.save(file_name)


    # Count positions where the channels differs
    cav1 = result_df.iloc[0]
    cav2 = result_df.iloc[1]

    differences_1_2 = 0
    different_aa_families_positions_1_2 = 0
    differences_1_3 = 0
    different_aa_families_positions_1_3 = 0
    differences_2_3 = 0
    different_aa_families_positions_2_3 = 0
    for column in result_df.columns:
        cav1_aa = cav1[column]
        cav2_aa = cav2[column]

        # Compare cav1 & cav2
        for aa in result_df[column][1:2]:  # Skip the first row
            if cav1_aa is not None and aa is not None and cav1_aa != aa:
                differences_1_2 += 1
                if result_df[column][3] is not None:
                    different_aa_families_positions_1_2 += 1

        # Compare cav1 & cav2
        for aa in result_df[column][2:3]:  # Skip the first row
            if cav1_aa is not None and aa is not None and cav1_aa != aa:
                differences_1_3 += 1
                if result_df[column][3] is not None:
                    different_aa_families_positions_1_3 += 1

        # Compare cav2 & cav3
        for aa in result_df[column][2:3]:  # Skip the first row
            if cav2_aa is not None and aa is not None and cav2_aa != aa:
                differences_2_3 += 1
                if result_df[column][3] is not None:
                    different_aa_families_positions_2_3 += 1


def find_mutant_segments(df):
    positions = sorted(df.columns)
    consecutive_mutations = []
    temp_segment = [positions[0]]

    for i in range(1, len(positions)):
        if positions[i] - positions[i - 1] == 1:
            temp_segment.append(positions[i])
        else:
            if len(temp_segment) >= 3:  # Save segments with 3 or more positions
                consecutive_mutations.append(temp_segment)
            temp_segment = [positions[i]]

    if len(temp_segment) >= 3:
        consecutive_mutations.append(temp_segment)

    for segment in consecutive_mutations:
        for pos in segment:
            df.loc["Consecutive Mutations", pos] = "*"

    return df
